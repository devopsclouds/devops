import pullrequestV2
from pullrequest import file
from pullrequest import username
from pullrequest import password



import json
import os
import glob
import requests
import csv
import datetime
import configparser
import pandas as pd
from openpyxl import load_workbook
import xlrd
import getpass
import sys








#os.chdir(r"../reports")
#file = input(" Please enter file name:")
####call file function. IT is defined function in fetch-pullrequest.py

print(file)
file_location = file  + ".xlsx"
print("File being created is :::" + file_location)


cp= configparser.RawConfigParser()
cp.read(r"./config/app.properties")

sys.dont_write_bytecode = True
####call the proxy function. function is defined in proxy.py
#callProxy()


parent_url = "https://jira.tools.telstra.com/browse/"



workbook = xlrd.open_workbook(file_location)

sheet = workbook.sheet_by_index(0)
nrows=sheet.nrows
#pull_request = []
jira = []
for row in range(nrows):
    
    jira.append(sheet.cell_value(row,2))
    #print(pull_request)
    #print(jira_issue)
  
issues = []
for issue in jira:
    if issue == 'jira_issue':
        print('')
    else:
        issues.append(issue)
res=[]    
for i in issues:
    k=i.split(', ')
    for j in k:
        if j=='':
            pass
        else:
            j.replace(' ',"")
            res.append(j)

resu = []
for x in res:
    if x not in resu:
                
        resu.append(x)

#cp= configparser.RawConfigParser()
#cp.read(r"app.properties")
jira_user = cp.get("auth", "jira_user")
jira_token = cp.get("auth", "jira_token")
headers = {"Content-Type": 'application/json'}
auth = (jira_user,jira_token) 


JiraRestApi= cp.get("DEFAULT","JiraRestApi")

defect_list = []
epic_list = []
issue_not_found = " issue is not found in jira"
issues_not_have_parent = " issue is not have a parent link"

#### To fetch the defect and epic details
for key_find in resu:
    defect_details = {}
    PIC_DETAILS ={}
    
    jira_rest_api_url = JiraRestApi + key_find
    #print(jira_rest_api_url)
    #print(jira_rest_api_url)
    responsed = requests.get(jira_rest_api_url, headers=headers, auth=auth)
    ##print(responsed.json())
    
    if "key" in responsed.json():
        if "fields" in responsed.json():
            issuetype = responsed.json()['fields']['issuetype']['name']
            ## print(issuetype)
            
            if issuetype == "Defect":
                if "key" in responsed.json():
                    if responsed.json()['key'] != None:
                        defect_details['defectid']=responsed.json()['key']
                    else:
                        defect_details['defectid']="None"
                else:
                    print('key not')
                if "name" in responsed.json()['fields']['issuetype']:
                    if responsed.json()['fields']['issuetype']['name'] != None:
                        defect_details['issuename']=responsed.json()['fields']['issuetype']['name']
                    else:
                        defect_details['issuename']="None"
                else:
                    print("key not found")
                if "name" in responsed.json()['fields']['priority']:
                    if responsed.json()['fields']['priority']['name'] != None:
                        defect_details['priority']=responsed.json()['fields']['priority']['name']
                    else:
                        #print("it is null")
                        defect_details['priority']="None"
                else:
                    print('not')
                if "components" in responsed.json()['fields']:
                    if responsed.json()['fields']['components'] != None:
                        compo = responsed.json()['fields']['components']
                        if compo ==[]:
                            defect_details['components']="None"
                        else:
                        
                            compos =[]
                            for i in range(len(compo)):
                                x=compo[i]
                                compos.append(x['name'])
                            if compos == []:
                                defect_details['components']="None"
                            else:
                                
                                defect_details['components']=compos
        
                    else:
                        defect_details['components']="None"
                else:
                    defect_details['components']="None"
        
                
                if "fixVersions" in responsed.json()['fields']:
                    
                    if responsed.json()['fields']['fixVersions'] != None:
                            fixe = responsed.json()['fields']['fixVersions']
                            if fixe ==[]:
                                defect_details['fixVersions']="None"
                                
                            else:
                                fixve = []
                                for i in range(len(fixe)):
                                    y=fixe[i]
                                    fixve.append(y['name'])
                                if fixve == []:
                                    defect_details['fixVersions']="None"
                                        
                                else:
                                    
                                    defect_details['fixVersions']=fixve
                                        
                    
                    else:
                            
                        defect_details['fixVersions']="None"
                else:
                    defect_details['fixVersions']="None"
        
                if "value" in responsed.json()['fields']['customfield_11454']:
                    if responsed.json()['fields']['customfield_11454']['value'] != None:
                        defect_details['FeatureTeam']=responsed.json()['fields']['customfield_11454']['value']
                    else:
                        defect_details['FeatureTeam']="None"
                else:
                    defect_details['FeatureTeam']="None"

                if "value" in responsed.json()['fields']['customfield_10750']:
                    if responsed.json()['fields']['customfield_10750']['value'] != None:
                        defect_details['DefectType']=responsed.json()['fields']['customfield_10750']['value']
                    else:
                        defect_details['DefectType']="None"
                else:
                    defect_details['DefectType']="None"
            
                if "value" in responsed.json()['fields']['customfield_10351']:
                    if responsed.json()['fields']['customfield_10351']['value'] != None:
                        defect_details['severity']=responsed.json()['fields']['customfield_10351']['value']
                    else:
                        defect_details['severity']="None"
                else:
                    defect_details['severity']="None"
            
                if "value" in responsed.json()['fields']['customfield_13955']:
                    if responsed.json()['fields']['customfield_13955']['value'] != None:
                        defect_details['environment']=responsed.json()['fields']['customfield_13955']['value']
                    else:
                        defect_details['environment']="None"
                else:
                    defect_details['environment']="None"
        
                if "value" in responsed.json()['fields']['customfield_15750']:
                    if responsed.json()['fields']['customfield_15750']['value'] != None:
                        defect_details['applicationtype']=responsed.json()['fields']['customfield_15750']['value']
                    else:
                        defect_details['applicationtype']="None"
                else:
                    defect_details['applicationtype']="None"
        
                if "summary" in responsed.json()['fields']['parent']['fields']:
                    if responsed.json()['fields']['parent']['fields']['summary'] != None:
                        defect_details['title']=responsed.json()['fields']['parent']['fields']['summary']
                    else:
                        defect_details['title']="None"
                else:
                    defect_details['title']="None"
            
                defect_list.append(defect_details)
                #print(defect_list)
            elif issuetype == "Service Request":
                print('Ignored this as it is service request :: ' + issuetype)
			
            else:
                if  responsed.json()['fields']['customfield_10152'] != None or responsed.json()['fields']['customfield_12954'] !=None:
                    
                    
                    if "key" in responsed.json():
                        PIC_DETAILS['issuetypeid']=responsed.json()['key']
                        ##print(PIC_DETAILS['issuetypeid'])
                        PIC_DETAILS['issuetypename']=responsed.json()['fields']['issuetype']['name']
                        if responsed.json()['fields']['issuetype']['name']=="Epic":
                            #print(PIC_DETAILS['issuetypename'])
                            if "name" in responsed.json()['fields']['priority']:
                                if "name" in responsed.json()['fields']['issuetype']:
                                    PIC_DETAILS['epic_priority']=responsed.json()['fields']['priority']['name']
                                    
                                else:
                                    PIC_DETAILS['epic_priority']="None" 
                                    
                            else:
                                PIC_DETAILS['epic_priority']="None"
                                
                            if "customfield_10153" in responsed.json()['fields']:
                                if responsed.json()['fields']['customfield_10153'] != None:
                                    PIC_DETAILS['epic_name']=responsed.json()['fields']['customfield_10153']
            
                                else:
                                #print("it is null")
                                    PIC_DETAILS['epic_name']="None"
                            else:
                                PIC_DETAILS['epic_name']="None"
                            if "name" in responsed.json()['fields']['status']:
                                if responsed.json()['fields']['status']['name'] != None:
                                    PIC_DETAILS['epic_status']=responsed.json()['fields']['status']['name']
            
                                else:
                                #print("it is null")
                                    PIC_DETAILS['epic_status']="None"
                            else:
                            

                               PIC_DETAILS['epic_status']="None"
                            if "fixVersions" in responsed.json()['fields']:
                                if responsed.json()['fields']['fixVersions'] != None:
                                    fix = responsed.json()['fields']['fixVersions']
                                    if fix ==[]:
                                        PIC_DETAILS['epic_fixVersions']="None"
                                    
                                    else:
                                        fixv = []
                                        for i in range(len(fix)):
                                            y=fix[i]
                                            fixv.append(y['name'])
                                        if fixv == []:
                                            PIC_DETAILS['epic_fixVersions']="None"
                                            #print(fixer)
                                        else:
                                            
                                            PIC_DETAILS['epic_fixVersions']=fixv
                                            
                        
                                else:
                                
                                    PIC_DETAILS['epic_fixVersions']="None"
                                
                            else:
                                PIC_DETAILS['epic_fixVersions']="None"
                                
                            if "components" in responsed.json()['fields']:
                                if responsed.json()['fields']['components'] != None:
                                    com = responsed.json()['fields']['components']
                                    if com ==[]:
                                        PIC_DETAILS['epic_components']="None"
                                    else:
                                
                                        comp =[]
                                        for i in range(len(com)):
                                            x=com[i]
                                            comp.append(x['name'])
                                        if comp == []:
                                            PIC_DETAILS['epic_components']="None"
                                        else:
                                            
                                            PIC_DETAILS['epic_components']=comp
                
                                else:
                                    PIC_DETAILS['epic_components']="None"
                            else:
                                PIC_DETAILS['epic_components']="None"
                            if "customfield_12954" in responsed.json()['fields']:
                                if responsed.json()['fields']['customfield_12954'] != None:
                                    PIC_DETAILS['epic_parentlink']= parent_url + str(responsed.json()['fields']['customfield_12954'])
                                else:
                                    PIC_DETAILS['epic_parentlink']="None"
                        
                            else:
                                PIC_DETAILS['epic_parentlink']="None"
                            if "value" in str(responsed.json()['fields']['customfield_12852']):
                                if str(responsed.json()['fields']['customfield_12852']['value']) != None:
                                    PIC_DETAILS['epic_ProgramIncrement']=str(responsed.json()['fields']['customfield_12852']['value'])
            
                                else:
                                    PIC_DETAILS['epic_ProgramIncrement']="None"
                            else:
                                PIC_DETAILS['epic_ProgramIncrement']="None"
                
                            if "customfield_10151" in responsed.json()['fields']:
                                if responsed.json()['fields']['customfield_10151'] != None:
                                    sprintname=""
                                    epicsprints= responsed.json()['fields']['customfield_10151']
                                    for i in range(len(epicsprints)):
                                        x=epicsprints[i]
                                        y=x.find('name')
                                        for j in range(y+5,len(x)):
                                            if x[j]==',':
                                                break
                                            sprintname+=x[j]
                                            #print(sprintname)
                                        
                                        
                                else:
                                    sprintname="None"
                            else:
                                sprintname="None"
        
                
                            if "value" in str(responsed.json()['fields']['customfield_11454']):
                                if str(responsed.json()['fields']['customfield_11454']['value'])!= None:
                                    PIC_DETAILS['epic_FeatureTeam']=str(responsed.json()['fields']['customfield_11454']['value'])
                                else:
                                    PIC_DETAILS['epic_FeatureTeam']="None"
                            else:
                                PIC_DETAILS['epic_FeatureTeam']="None"
                            
                            PIC_DETAILS['epic_sprint']=sprintname    
                            
                        else:
                            epicid = responsed.json()['fields']['customfield_10152']
                        
                            
               
                        
                            jira_rest_api_url4 = JiraRestApi + epicid
                            
                            respons = requests.get(jira_rest_api_url4, headers=headers, auth=auth)
                            #values2 = respons.json()['fields']
                            #json_str2 = json.dumps(values2)
                            #json_list2 = json.loads(json_str2)
                            #type(json_list2)
                                
                            
        
                            if "key" in respons.json():
                                epickey=respons.json()['key']
                            if "name" in respons.json()['fields']['issuetype']:
                                if respons.json()['fields']['issuetype']['name'] != None:
                                    epicissuetype =respons.json()['fields']['issuetype']['name']
                                else:
                                    epicissuetype="None"
                            else:
                                print('name not found')
                            
                            if "name" in respons.json()['fields']['priority']:
                                if respons.json()['fields']['priority']['name'] != None:
                                    epicpriority=respons.json()['fields']['priority']['name']
                
                                else:
                                    #print("it is null")
                                    epicpriority="None"
                            else:
                                print('priority not found')
                            if "customfield_10153" in respons.json()['fields']:
                                if respons.json()['fields']['customfield_10153'] != None:
                                    epicname=respons.json()['fields']['customfield_10153']
                
                                else:
                                    #print("it is null")
                                    epicname="None"
                            else:
                                epicname="None"
                            if "name" in respons.json()['fields']['status']:
                                if respons.json()['fields']['status']['name'] != None:
                                    epicstatus=respons.json()['fields']['status']['name']
                
                                else:
                                    #print("it is null")
                                    epicstatus="None"
                            else:
                                #print("not found")

                                epicstatus="None"
                            if "fixVersions" in respons.json()['fields']:
                                if respons.json()['fields']['fixVersions'] != None:
                                    fix = respons.json()['fields']['fixVersions']
                                    if fix ==[]:
                                        fixer="None"
                                        
                                    else:
                                        fixv = []
                                        for i in range(len(fix)):
                                            y=fix[i]
                                            fixv.append(y['name'])
                                            if fixv == []:
                                                fixer="None"
                                                print(fixer)
                                            else:
                                                
                                                fixer=fixv
                                                
                            
                                else:
                                    
                                    fixer="None"
                                    
                            else:
                               fixer="None"
                               print(fixer)
                            if "components" in respons.json()['fields']:
                                if respons.json()['fields']['components'] != None:
                                    com = respons.json()['fields']['components']
                                    if com ==[]:
                                        st="None"
                                    else:
                                    
                                        comp =[]
                                        for i in range(len(com)):
                                            x=com[i]
                                            comp.append(x['name'])
                                        if comp == []:
                                            st="None"
                                        else:
                                            
                                            st=comp
                    
                                else:
                                    st="None"
                            else:
                                st="None"
                            if "customfield_12954" in respons.json()['fields']:
                                if respons.json()['fields']['customfield_12954'] != None:
                                    epicparentlink= parent_url + str(respons.json()['fields']['customfield_12954'])
                                else:
                                    epicparentlink="None"
                            
                            else:
                                epicparentlink="None"
                            if "value" in str(respons.json()['fields']['customfield_12852']):
                                if str(respons.json()['fields']['customfield_12852']['value']) != None:
                                    epicprogramincrement= str(respons.json()['fields']['customfield_12852']['value'])
                
                                else:
                                    epicprogramincrement="None"
                            else:
                                epicprogramincrement="None"
                    
                            if "customfield_10151" in respons.json()['fields']:
                                if respons.json()['fields']['customfield_10151'] != None:
                                    sprintname=""
                                    epicsprints= respons.json()['fields']['customfield_10151']
                                    for i in range(len(epicsprints)):
                                        x=epicsprints[i]
                                        y=x.find('name')
                                        for j in range(y+5,len(x)):
                                            if x[j]==',':
                                                break
                                            sprintname+=x[j]
                                            
                                            
                                else:
                                    sprintname="None"
                            else:
                                sprintname="None"
            
                    
                            if "value" in str(respons.json()['fields']['customfield_11454']):
                                if str(respons.json()['fields']['customfield_11454']['value'])!= None:
                                    epicfeatureteam=str(respons.json()['fields']['customfield_11454']['value'])
                                else:
                                    epicfeatureteam="None"
                            else:
                                epicfeatureteam="None"
                            PIC_DETAILS['epic_id']=epickey
                            PIC_DETAILS['epic_issuename']=epicissuetype
                            PIC_DETAILS['epic_priority']=epicpriority
                            PIC_DETAILS['epic_name']=epicname
                            PIC_DETAILS['epic_status']=epicstatus
                            PIC_DETAILS['epic_components']=st
                            PIC_DETAILS['epic_fixVersions']=fixer
                            PIC_DETAILS['epic_parentlink']=epicparentlink
                            PIC_DETAILS['epic_ProgramIncrement']=epicprogramincrement
                            PIC_DETAILS['epic_sprint']=sprintname
                            PIC_DETAILS['epic_FeatureTeam']=epicfeatureteam
                    else:
                        print('key not found')
                        
                        
                else:
                    issue_not_have_parent = key_find + "::: issue has no parent"
                    print(issue_not_have_parent)
                    PIC_DETAILS['issuetypeid'] = key_find
                    PIC_DETAILS['issuetypename'] = responsed.json()['fields']['issuetype']['name']
                    PIC_DETAILS['epic_id']="No parent epic details"
                    PIC_DETAILS['epic_issuename']="No parent epic details"
                    PIC_DETAILS['epic_priority']="No parent epic details"
                    PIC_DETAILS['epic_name']="No parent epic details"
                    PIC_DETAILS['epic_status']="No parent epic details"
                    PIC_DETAILS['epic_components']="No parent epic details"
                    PIC_DETAILS['epic_fixVersions']="No parent epic details"
                    
                    PIC_DETAILS['epic_parentlink']="No parent epic details"
                    PIC_DETAILS['epic_ProgramIncrement']="No parent epic details"
                    PIC_DETAILS['epic_sprint']="No parent epic details"
                    PIC_DETAILS['epic_FeatureTeam']="No parent epic details"
                    
                epic_list.append(PIC_DETAILS)
        else:
            print('fields not found')

    else:
        issue_is_not_found_in_jira = key_find + issue_not_found
        print(issue_is_not_found_in_jira)

##### To create report for epic and defect
book = load_workbook(file_location)
writer = pd.ExcelWriter(file_location, engine = 'openpyxl')
writer.book = book
df5 = pd.DataFrame(epic_list,columns=['issuetypeid','issuetypename','epic_id','epic_name','epic_priority','epic_status','epic_components','epic_fixVersions','epic_parentlink','epic_ProgramIncrement','epic_sprint','epic_FeatureTeam','epic_issuename'])
df6 = pd.DataFrame(defect_list,columns=['defectid','issuename','priority','components','fixVersions','FeatureTeam','DefectType','severity','environment','applicationtype','title'])

 

df5.to_excel(writer, sheet_name = 'Epic')
df6.to_excel(writer, sheet_name = 'Defect')
writer.save()
writer.close()


book1 = load_workbook(file_location)
sheet1 = book1['Epic']
sheet1.column_dimensions['A'].width = 0
sheet1.column_dimensions['B'].width = 50
sheet1.column_dimensions['C'].width = 50
sheet1.column_dimensions['D'].width = 50
sheet1.column_dimensions['E'].width = 70
sheet1.column_dimensions['F'].width = 60
sheet1.column_dimensions['G'].width = 50
sheet1.column_dimensions['H'].width = 50
sheet1.column_dimensions['I'].width = 50
sheet1.column_dimensions['J'].width = 50
sheet1.column_dimensions['k'].width = 50
sheet1.column_dimensions['L'].width = 70
sheet1.column_dimensions['M'].width = 50
sheet1.column_dimensions['N'].width = 50

book1.save(file_location)
book1.close()



book2 = load_workbook(file_location)
sheet2 = book2['Defect']
sheet2.column_dimensions['A'].width = 0
sheet2.column_dimensions['B'].width = 30
sheet2.column_dimensions['C'].width = 30
sheet2.column_dimensions['D'].width = 30
sheet2.column_dimensions['E'].width = 30
sheet2.column_dimensions['F'].width = 30
sheet2.column_dimensions['G'].width = 30
sheet2.column_dimensions['H'].width = 30
sheet2.column_dimensions['I'].width = 30

sheet2.column_dimensions['J'].width = 30
sheet2.column_dimensions['K'].width = 30
sheet2.column_dimensions['L'].width = 100
book2.save(file_location)
book2.close()

cwd = os.getcwd()
print(cwd)